#!/usr/bin/env python3
"""
Minimal dataset validator for prompt-optimizer.

Default behavior:
- Scan a directory for dataset files: *.jsonl, *.xlsx, *.xlsm
- Validate each dataset file contains `ground_truth`:
  - JSONL: every non-empty line must be a JSON object and include key `ground_truth`
  - Excel: header row must include column `ground_truth` (requires openpyxl)
- Infer dynamic input keys: union of all fields/columns excluding `ground_truth`

Outputs:
- Prints a short summary to stdout
- Optionally writes a JSON report (or writes to $EVO_OUTPUT_DIR if set)
"""

from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple


def _iter_jsonl_objects(path: Path) -> Iterable[Tuple[int, Dict[str, Any]]]:
    with path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            raw = line.strip()
            if not raw:
                continue
            obj = json.loads(raw)
            if not isinstance(obj, dict):
                raise ValueError(f"{path.name}:{line_no} is not a JSON object")
            yield line_no, obj


def _validate_jsonl(path: Path) -> Tuple[int, Set[str]]:
    rows = 0
    keys: Set[str] = set()
    for line_no, obj in _iter_jsonl_objects(path):
        rows += 1
        if "ground_truth" not in obj:
            raise ValueError(f"{path.name}:{line_no} missing required field: ground_truth")
        keys.update(k for k in obj.keys() if k != "ground_truth")
    if rows == 0:
        raise ValueError(f"{path.name} has no records")
    return rows, keys


def _validate_excel(path: Path) -> Tuple[int, Set[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise ValueError("Excel format requires openpyxl (pip install openpyxl)") from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ValueError(f"{path.name} is empty (no header row)")

    cols: List[str] = []
    for cell in header:
        cols.append("" if cell is None else str(cell).strip())

    if "ground_truth" not in cols:
        raise ValueError(f"{path.name} missing required column: ground_truth")

    dynamic = {c for c in cols if c and c != "ground_truth"}
    row_count = 0
    for row in rows_iter:
        if row is None:
            continue
        row_count += 1
    if row_count == 0:
        raise ValueError(f"{path.name} has no data rows")
    return row_count, dynamic


def _discover_dataset_files(input_dir: Path) -> List[Path]:
    files: List[Path] = []
    for p in input_dir.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() in {".jsonl", ".xlsx", ".xlsm"}:
            files.append(p)
    return sorted(files)


def _write_report(report: Dict[str, Any], report_path: Optional[Path]) -> Optional[Path]:
    if report_path:
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        return report_path

    evo_out = os.getenv("EVO_OUTPUT_DIR")
    if evo_out:
        out_path = Path(evo_out).expanduser().resolve() / "dataset_validation_report.json"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        return out_path

    return None


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate dataset files contain ground_truth (jsonl/excel).")
    parser.add_argument("--dir", required=True, help="Directory containing dataset files (*.jsonl/*.xlsx/*.xlsm)")
    parser.add_argument("--report-json", default=None, help="Write report JSON to this path")
    args = parser.parse_args()

    input_dir = Path(args.dir).expanduser().resolve()
    if not input_dir.exists() or not input_dir.is_dir():
        print(f"❌ Invalid --dir (not a directory): {input_dir}")
        return 1

    files = _discover_dataset_files(input_dir)
    if not files:
        print(f"❌ No dataset files found in {input_dir} (expected *.jsonl/*.xlsx/*.xlsm)")
        return 1

    per_file: List[Dict[str, Any]] = []
    dynamic_keys: Set[str] = set()

    for path in files:
        try:
            if path.suffix.lower() == ".jsonl":
                rows, keys = _validate_jsonl(path)
                fmt = "jsonl"
            else:
                rows, keys = _validate_excel(path)
                fmt = "excel"
            dynamic_keys |= keys
            per_file.append({"file": path.name, "format": fmt, "rows": rows, "ok": True})
        except Exception as exc:
            per_file.append({"file": path.name, "ok": False, "error": str(exc)})

    ok = all(item.get("ok") is True for item in per_file)
    report = {
        "ok": ok,
        "input_dir": str(input_dir),
        "files": per_file,
        "dynamic_input_keys": sorted(dynamic_keys),
    }

    written = _write_report(report, Path(args.report_json).expanduser().resolve() if args.report_json else None)

    if ok:
        print(f"✅ dataset validation passed ({len(files)} files). dynamic_input_keys={report['dynamic_input_keys']}")
        if written:
            print(f"✅ report: {written}")
        return 0

    print("❌ dataset validation failed:")
    for item in per_file:
        if not item.get("ok"):
            print(f" - {item['file']}: {item.get('error')}")
    if written:
        print(f"⚠️  report: {written}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())

